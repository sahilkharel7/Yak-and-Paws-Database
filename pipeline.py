import sys
import time

import requests
import zipcodes
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os


# All 50 US states + DC, mapping names and abbreviations for flexible input.
US_STATES = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT", "delaware": "DE",
    "florida": "FL", "georgia": "GA", "hawaii": "HI", "idaho": "ID",
    "illinois": "IL", "indiana": "IN", "iowa": "IA", "kansas": "KS",
    "kentucky": "KY", "louisiana": "LA", "maine": "ME", "maryland": "MD",
    "massachusetts": "MA", "michigan": "MI", "minnesota": "MN",
    "mississippi": "MS", "missouri": "MO", "montana": "MT", "nebraska": "NE",
    "nevada": "NV", "new hampshire": "NH", "new jersey": "NJ",
    "new mexico": "NM", "new york": "NY", "north carolina": "NC",
    "north dakota": "ND", "ohio": "OH", "oklahoma": "OK", "oregon": "OR",
    "pennsylvania": "PA", "rhode island": "RI", "south carolina": "SC",
    "south dakota": "SD", "tennessee": "TN", "texas": "TX", "utah": "UT",
    "vermont": "VT", "virginia": "VA", "washington": "WA",
    "west virginia": "WV", "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}

# Reverse lookup: abbreviation -> full name
ABBR_TO_NAME = {v: k.title() for k, v in US_STATES.items()}


# ---------------------------------------------------------------------------
# Validates a state name or abbreviation. Returns (full_name, abbreviation)
# or None if the input is not a recognized US state.
# ---------------------------------------------------------------------------
def validate_state(state_input):
    cleaned = state_input.strip().lower()

    # Check if it's a full state name
    if cleaned in US_STATES:
        abbr = US_STATES[cleaned]
        return (cleaned.title(), abbr)

    # Check if it's a 2-letter abbreviation
    upper = cleaned.upper()
    if upper in ABBR_TO_NAME:
        return (ABBR_TO_NAME[upper], upper)

    return None


# ---------------------------------------------------------------------------
# Returns all ZIP codes for a state with valid coordinates.
# Uses the zipcodes library (offline local database, no API calls).
# Returns a list of (zip_code, lat, lng) tuples.
# ---------------------------------------------------------------------------
def get_state_zip_codes(state_abbr):
    all_zips = zipcodes.filter_by(state=state_abbr, zip_code_type="STANDARD")
    result = []
    for z in all_zips:
        if not z.get("active"):
            continue
        lat_str = z.get("lat", "0.0000")
        lng_str = z.get("long", "0.0000")
        try:
            lat = float(lat_str)
            lng = float(lng_str)
        except (ValueError, TypeError):
            continue
        # Skip entries with zero/missing coordinates
        if lat == 0.0 and lng == 0.0:
            continue
        result.append((z["zip_code"], lat, lng))
    return result


# ---------------------------------------------------------------------------
# Searches for nearby places using the Google Places Nearby Search API.
# Handles pagination via next_page_token (up to 3 pages max from Google).
# Returns a list of raw place result dictionaries.
# ---------------------------------------------------------------------------
def search_nearby_places(lat, lng, search_term, api_key):
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": 5000,
        "keyword": search_term,
        "key": api_key,
    }

    all_results = []

    try:
        while True:
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            data = response.json()

            if data["status"] not in ("OK", "ZERO_RESULTS"):
                break

            all_results.extend(data.get("results", []))

            next_token = data.get("next_page_token")
            if not next_token:
                break

            # Google requires a short delay before the token becomes valid
            time.sleep(2)
            params = {"pagetoken": next_token, "key": api_key}

        return all_results

    except requests.RequestException:
        return all_results


# ---------------------------------------------------------------------------
# Fetches full details for a single place using its place_id.
# Returns the detail dictionary, or None on failure.
# ---------------------------------------------------------------------------
def get_place_details(place_id, api_key):
    url = "https://maps.googleapis.com/maps/api/place/details/json"
    fields = (
        "name,place_id,formatted_address,address_components,"
        "formatted_phone_number,international_phone_number,website,url,"
        "rating,user_ratings_total,price_level,business_status,types,"
        "geometry,opening_hours"
    )
    params = {"place_id": place_id, "fields": fields, "key": api_key}

    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()

        if data["status"] != "OK":
            return None

        return data.get("result")

    except requests.RequestException:
        return None


# ---------------------------------------------------------------------------
# Removes duplicate places based on place_id. Returns the deduplicated list.
# ---------------------------------------------------------------------------
def deduplicate_places(places_list):
    seen = set()
    unique = []
    for place in places_list:
        pid = place.get("place_id")
        if pid and pid not in seen:
            seen.add(pid)
            unique.append(place)
    return unique


# ---------------------------------------------------------------------------
# Helper: extracts a specific address component type from the
# address_components list returned by the Places API.
# ---------------------------------------------------------------------------
def _extract_component(components, target_type):
    for comp in components:
        if target_type in comp.get("types", []):
            return comp.get("long_name", "")
    return ""


# ---------------------------------------------------------------------------
# Helper: extracts the short_name for state from address_components.
# ---------------------------------------------------------------------------
def _extract_state(components):
    for comp in components:
        if "administrative_area_level_1" in comp.get("types", []):
            return comp.get("short_name", "")
    return ""


# ---------------------------------------------------------------------------
# Builds a flat record dictionary from a place details response.
# source_zip tracks which ZIP code search discovered this business.
# Missing fields default to an empty string.
# ---------------------------------------------------------------------------
def build_record(place_details, search_term, source_zip):
    if not place_details:
        return None

    components = place_details.get("address_components", [])
    geometry = place_details.get("geometry", {})
    location = geometry.get("location", {})
    opening_hours = place_details.get("opening_hours", {})
    periods_text = opening_hours.get("weekday_text", [])

    # Map each day name to its hours string from weekday_text
    day_hours = {
        "Monday": "", "Tuesday": "", "Wednesday": "", "Thursday": "",
        "Friday": "", "Saturday": "", "Sunday": "",
    }
    for entry in periods_text:
        for day in day_hours:
            if entry.startswith(day):
                day_hours[day] = entry.split(": ", 1)[1] if ": " in entry else entry

    types_list = place_details.get("types", [])

    return {
        "business_name": place_details.get("name", ""),
        "place_id": place_details.get("place_id", ""),
        "formatted_address": place_details.get("formatted_address", ""),
        "street_number": _extract_component(components, "street_number"),
        "street_name": _extract_component(components, "route"),
        "city": _extract_component(components, "locality"),
        "state": _extract_state(components),
        "zip_code": _extract_component(components, "postal_code"),
        "country": _extract_component(components, "country"),
        "latitude": location.get("lat", ""),
        "longitude": location.get("lng", ""),
        "phone_local": place_details.get("formatted_phone_number", ""),
        "phone_international": place_details.get("international_phone_number", ""),
        "website": place_details.get("website", ""),
        "google_maps_url": place_details.get("url", ""),
        "rating": place_details.get("rating", ""),
        "user_ratings_total": place_details.get("user_ratings_total", ""),
        "price_level": place_details.get("price_level", ""),
        "business_status": place_details.get("business_status", ""),
        "business_types": ", ".join(types_list),
        "open_now": opening_hours.get("open_now", ""),
        "hours_monday": day_hours["Monday"],
        "hours_tuesday": day_hours["Tuesday"],
        "hours_wednesday": day_hours["Wednesday"],
        "hours_thursday": day_hours["Thursday"],
        "hours_friday": day_hours["Friday"],
        "hours_saturday": day_hours["Saturday"],
        "hours_sunday": day_hours["Sunday"],
        "search_term": search_term,
        "source_zip": source_zip,
    }


# ---------------------------------------------------------------------------
# Saves a list of record dictionaries to a formatted Excel file.
# Returns the filename that was saved.
# ---------------------------------------------------------------------------
def save_to_excel(records, state_abbr):
    filename = f"pet_stores_{state_abbr}.xlsx"

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pet Businesses"

        if not records:
            wb.save(filename)
            return filename

        headers = list(records[0].keys())

        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        for row_idx, record in enumerate(records, start=2):
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(header, ""))

        # Auto-size columns based on content width (capped at 50)
        for col_idx, header in enumerate(headers, start=1):
            max_length = len(str(header))
            for row_idx in range(2, len(records) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        wb.save(filename)
        print(f"\n[OK] Saved to {filename}")
        return filename

    except Exception as e:
        print(f"[ERROR] Failed to save Excel file: {e}")
        return None


# ---------------------------------------------------------------------------
# Main pipeline: searches every ZIP code in a state for pet businesses.
# ---------------------------------------------------------------------------
def main():
    # Step 1 — Get state from user
    state_input = input("Enter state name or abbreviation: ").strip()

    state_info = validate_state(state_input)
    if state_info is None:
        print("[ERROR] Invalid state. Please enter a US state name or 2-letter abbreviation.")
        sys.exit(1)

    state_name, state_abbr = state_info
    print(f"\nState: {state_name} ({state_abbr})")

    # Step 2 — Load API key from .env
    load_dotenv()
    api_key = os.getenv("GOOGLE_PLACES_API_KEY")

    if not api_key:
        print("[ERROR] GOOGLE_PLACES_API_KEY not found in .env file.")
        sys.exit(1)

    # Step 3 — Get all ZIP codes for the state (local database, no API calls)
    zip_codes = get_state_zip_codes(state_abbr)

    if not zip_codes:
        print(f"[ERROR] No ZIP codes found for {state_name}.")
        sys.exit(1)

    print(f"Found {len(zip_codes)} ZIP codes with valid coordinates")

    # Step 4 — Search every ZIP code with 3 search terms
    search_terms = ["pet store", "pet supplies", "pet shop"]
    all_raw_results = []
    # Track which ZIP each raw result came from
    result_source_zips = {}
    total_zips = len(zip_codes)

    print(f"\nSearching {total_zips} ZIP codes (3 search terms each)...\n")

    for zip_idx, (zip_code, lat, lng) in enumerate(zip_codes, start=1):
        zip_count = 0
        for term in search_terms:
            results = search_nearby_places(lat, lng, term, api_key)
            for r in results:
                pid = r.get("place_id")
                if pid and pid not in result_source_zips:
                    result_source_zips[pid] = zip_code
            zip_count += len(results)
            all_raw_results.extend(results)
            # Small delay between API calls to avoid rate limiting
            time.sleep(0.1)

        print(f"  [{zip_idx}/{total_zips}] ZIP {zip_code} — {zip_count} results")

    raw_count = len(all_raw_results)
    print(f"\nTotal raw results: {raw_count}")

    if raw_count == 0:
        print(f"[INFO] No pet businesses found in {state_name}. Exiting.")
        sys.exit(0)

    # Step 5 — Deduplicate across all ZIP code searches
    unique_places = deduplicate_places(all_raw_results)
    dedup_count = len(unique_places)
    print(f"After deduplication: {dedup_count}")

    # Step 6 — Get full details for each unique place
    print(f"\nFetching details for {dedup_count} unique businesses...\n")
    detailed_places = []
    for i, place in enumerate(unique_places, start=1):
        pid = place.get("place_id")
        name = place.get("name", "Unknown")
        print(f"  [{i}/{dedup_count}] {name}")
        details = get_place_details(pid, api_key)
        if details:
            # Attach the source ZIP that first found this place
            details["_source_zip"] = result_source_zips.get(pid, "")
            detailed_places.append(details)
        time.sleep(0.1)

    # Step 7 — Build clean records
    records = []
    for details in detailed_places:
        source_zip = details.pop("_source_zip", "")
        record = build_record(details, "pet store", source_zip)
        if record:
            records.append(record)

    if not records:
        print(f"[INFO] No records could be built for {state_name}. Exiting.")
        sys.exit(0)

    # Step 8 — Save to Excel
    filename = save_to_excel(records, state_abbr)

    # Step 9 — Print summary
    print("\n" + "=" * 50)
    print("PIPELINE SUMMARY")
    print("=" * 50)
    print(f"  State:                     {state_name} ({state_abbr})")
    print(f"  ZIP codes searched:        {total_zips}")
    print(f"  Raw results found:         {raw_count}")
    print(f"  After deduplication:        {dedup_count}")
    print(f"  Records saved:             {len(records)}")
    if filename:
        print(f"  Output file:               {filename}")
    print("=" * 50)


if __name__ == "__main__":
    main()
